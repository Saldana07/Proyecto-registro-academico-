from audioop import reverse
from django import forms
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
import openpyxl


from django.contrib.auth.models import User
# Create your views here.
from django.contrib import auth
from django.shortcuts import render, redirect
from django.contrib import auth
from .forms import DisponibilidadForm, LoginForm, registro_form,EditUserForm,ProyeccionForm,RestringirFechasForm,EditDisponibilidadForm,CronogramaForm,EditCronogramaForm,EditCronogramaForm1
from django.contrib.auth import login, authenticate
from django.shortcuts import render, get_object_or_404
from .models import Disponibilidad, Proyeccion,Asignatura,Mensaje,Restriccion,Programacion, Salones
from django.contrib.auth.models import Group, Permission
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from .models import Programacion, Asignatura,Programas,Cronograma


# Create your views here.
def inicio_view(request):
    return render(request,'base.html')

def profesor_view(request):
    return render(request,'baseProfesor.html')


def coordinador_view(request):
    return render(request,'baseCoordinacion.html')


def proyeccion_view(request):
    return render(request,'proyeccion.html')

def asignatura_view(request):
    return render(request,'asignatura.html')

def desactivar_view(request):
    return render(request,'DesactivarUsuario.html')





def usuarios_view(request):
    usuarios= User.objects.all()
    user= User.objects.all().prefetch_related('grups')
    context = {
        'usuarios':usuarios,
        'user':user,
        
    }
    return render(request,'usuarios.html',context)





def login_view(request):
    if request.method == 'POST':
        form = LoginForm(data=request.POST)
        if form.is_valid():
            user = auth.authenticate(username=form.cleaned_data.get('username'),
                                      password=form.cleaned_data.get('password'))
            auth.login(request, user)
            return redirect('/inicio/')
  
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})


def logout_view(request):
    return redirect('/')  # Cambia 'login' por la URL de tu página de inicio de sesión

def registro_view(request):
    if request.method == 'POST':
        form = registro_form(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = form.cleaned_data['username']
            user.save()
            group = form.cleaned_data.get('group')
            if group:
                group.user_set.add(user)
            
            return redirect('/inicio/registro/')
    else:
        form = registro_form()
    return render(request, 'registro.html', {'form': form})





def edit_user(request):
    user = request.user
    group = user.groups.first()
    form = EditUserForm(request.POST or None, instance=user,initial={'group': group})
    if form.is_valid():
        form.save()
        return redirect('/inicio/editar_usuario')
    return render(request, 'editarusuario.html', {'form': form})



def desactivar_usuario(request, email):
    
    usuario = get_object_or_404(User, email=email)
    usuario.is_active = not usuario.is_active  # invertir el estado actual
    usuario.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))


def activarusuarios_view(request):
    usuarios= User.objects.all()
    user= User.objects.all().prefetch_related('grups')
    
    context = {
        'usuarios':usuarios,
        'user':user,
        
    }
    return render(request,'activarDesactivar.html',context)


def cargar_usuarios(request):
    if request.method == 'POST' and request.FILES['ejemplo']:
        # Obtener el archivo del formulario
        file = request.FILES['ejemplo']
        # Abrir el archivo con openpyxl
        workbook = openpyxl.load_workbook(file)
        # Seleccionar la hoja que contiene los datos de usuario
        sheet = workbook['Sheet1']
        # Iterar sobre cada fila de la hoja y guardar los datos en la base de datos
        for row in sheet.iter_rows(min_row=2, values_only=True):
            username, email, password, first_name, last_name, group_name = row
            password = str(password) # Convertir la contraseña en una cadena
            # Obtener o crear el usuario
            user, created = User.objects.get_or_create(username=username, email=email)
            user.first_name = first_name
            user.last_name = last_name
            user.set_password(password)
            user.save()
            # Obtener o crear el grupo
            group, created = Group.objects.get_or_create(name=group_name)
            # Agregar al usuario al grupo
            user.groups.add(group)
        # Redirigir a una página de éxito después de guardar los datos
        return render(request, 'cargar.html')
    else:
        # Renderizar el formulario para cargar el archivo si el método no es POST
        return render(request, 'cargar.html')


def enviar_mensaje(request):
    if request.method == 'POST':
        receptor_id = request.POST['receptor']
        receptor = User.objects.get(id=receptor_id)
        contenido = request.POST['contenido']
        archivo_adjunto = request.FILES.get('archivo_adjunto')  # Obtener el archivo adjunto

        mensaje = Mensaje(emisor=request.user, receptor=receptor, contenido=contenido)
        mensaje.archivo_adjunto = archivo_adjunto  # Asignar el archivo adjunto al mensaje
        mensaje.save()

        # Puedes agregar lógica adicional aquí, como notificar al receptor por correo electrónico

    usuarios = User.objects.exclude(id=request.user.id)  # Obtén todos los usuarios excepto el usuario actual
    return render(request, 'enviar_mensaje.html', {'usuarios': usuarios})



def buzon_mensajes(request):
    mensajes_recibidos = Mensaje.objects.filter(receptor=request.user).order_by('-fecha_envio')
    return render(request, 'buzon_mensajes.html', {'mensajes_recibidos': mensajes_recibidos})

def get_datos():
    proyecciones = Proyeccion.objects.select_related('id_programas', 'id_asignatura')
    datos = []

    for proyeccion in proyecciones:
        semestre = proyeccion.semestre
        jornada = proyeccion.id_programas.jornada
        codigo = proyeccion.id_asignatura.codigo
        asignatura = proyeccion.id_asignatura.nombre
        creditos = proyeccion.id_asignatura.creditos
        intensidad = proyeccion.id_asignatura.intensidad
        total_semana = 19  # Aquí podrías definir la cantidad de semanas que desees
        num_profesores = 1  # Aquí podrías definir el número de profesores que desees
        total_a_pagar = intensidad * total_semana * num_profesores

        datos.append({
            'semestre': semestre,
            'jornada': jornada,
            'codigo': codigo,
            'asignatura': asignatura,
            'creditos': creditos,
            'intensidad': intensidad,
            'total_semana': total_semana,
            'num_profesores': num_profesores,
            'total_a_pagar': total_a_pagar,
        })

    return datos

def proyeccion_view(request):
    if request.method == 'POST':
        form = ProyeccionForm(request.POST)
        if form.is_valid():
            id_programas = form.cleaned_data['id_programas']
            id_asignatura = form.cleaned_data['id_asignatura']
            semestre = form.cleaned_data['semestre']
            
            proyeccion = Proyeccion.objects.create(id_programas=id_programas, id_asignatura=id_asignatura,semestre=semestre)
            return redirect('/inicio/proyeccion')
    else:
        form = ProyeccionForm()
        reporte_context = reporte_view(request)
        context = {
        'form': form,
        'datos': reporte_context,
        }
     

    return render(request, 'proyeccion.html',context)

def descargar_tabla(request):
    datos = get_datos() # Suponiendo que tienes una función que te devuelve los datos de la tabla
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="proyeccion.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar encabezados a la hoja de cálculo
    sheet['A1'] = 'Semestre'
    sheet['B1'] = 'Jornada'
    sheet['C1'] = 'Código'
    sheet['D1'] = 'Asignatura'
    sheet['E1'] = 'Créditos'
    sheet['F1'] = 'Intensidad'
    sheet['G1'] = 'Total Semana'
    sheet['H1'] = 'N° Profesores'
    sheet['I1'] = 'Total a Pagar'

    # Agregar datos a la hoja de cálculo
    row = 2
    for dato in datos:
        sheet.cell(row=row, column=1).value = dato['semestre']
        sheet.cell(row=row, column=2).value = dato['jornada']
        sheet.cell(row=row, column=3).value = dato['codigo']
        sheet.cell(row=row, column=4).value = dato['asignatura']
        sheet.cell(row=row, column=5).value = dato['creditos']
        sheet.cell(row=row, column=6).value = dato['intensidad']
        sheet.cell(row=row, column=7).value = dato['total_semana']
        sheet.cell(row=row, column=8).value = dato['num_profesores']
        sheet.cell(row=row, column=9).value = dato['total_a_pagar']
        row += 1

    workbook.save(response)
    return response


def toggle_permission(request, group_id, permission_name):
    group = Group.objects.get(id=group_id)
    permission = Permission.objects.get(codename=permission_name)

    if 'add' in request.GET:
        group.permissions.add(permission)
    elif 'remove' in request.GET:
        group.permissions.remove(permission)

    return redirect('group_detail', group_id=group.id)



def reporte_view(request):

    proyecciones = Proyeccion.objects.select_related('id_programas', 'id_asignatura')
    datos = []


    total_intensidad = 0
    total_semana = 0
    total_profesores = 0
    total_pago = 0

    for proyeccion in proyecciones:
        semestre = proyeccion.semestre
        jornada = proyeccion.id_programas.jornada
        codigo = proyeccion.id_asignatura.codigo
        asignatura = proyeccion.id_asignatura.nombre
        creditos = proyeccion.id_asignatura.creditos
        intensidad = proyeccion.id_asignatura.intensidad
        total_semana = proyeccion.total_semana
        num_profesores = proyeccion.num_profesores

        datos.append({
            'semestre':semestre,
            'jornada': jornada,
            'codigo': codigo,
            'asignatura': asignatura,
            'creditos': creditos,
            'intensidad': intensidad,
            'total_semana': total_semana,
            'num_profesores': num_profesores,
            'total_a_pagar': intensidad*total_semana,
        })
        total_intensidad+= intensidad
        total_semana += 19
        total_profesores += 1
        total_pago += intensidad*19

    context = {
    
        'datos': datos,
        'total_intensidad': total_intensidad,
        'total_semana': total_semana,
        'total_profesores': total_profesores,
        'total_pago': total_pago,
    }
   
    return render(request, 'reporte.html', context)    



def proyeccion_list(request):
    proyecciones = Proyeccion.objects.all()
    return render(request, 'editarProyeccion.html', {'proyecciones': proyecciones})


def editar_proyeccion(request, id):
    proyeccion = get_object_or_404(Proyeccion, id=id)
    if request.method == 'POST':
        form = ProyeccionForm(request.POST, instance=proyeccion)
        if form.is_valid():
            form.save()
            return redirect('/inicio/editar')
    else:
        form = ProyeccionForm(instance=proyeccion)

    return render(request, 'editarProyeccion.html', {'form': form})


from django.contrib.auth.decorators import login_required, user_passes_test


def disponibilidad(request):
    ordenar_por = request.GET.get('ordenar_por', 'fecha') # por defecto, ordenar por fecha
    disponibilidad = Disponibilidad.objects.order_by(ordenar_por)
    context = {
        'disponibilidad': disponibilidad,
        'ordenar_por': ordenar_por,
    }
    return render(request, 'verDispo.html', context)





def restringir(request):
    fecha_actual = datetime.datetime.now().strftime('%Y-%m-%d')
    
    if request.method == 'POST':
        if 'cancelar_restriccion' in request.POST:
            Restriccion.objects.all().delete()  # eliminar todas las restricciones
            messages.success(request, 'Se ha cancelado la restricción de fechas.')
            return redirect('/inicio/seleccionar-fechas')
        else:
            fecha_inicio = datetime.datetime.strptime(request.POST.get('fecha_inicio'), '%Y-%m-%d')
            fecha_fin = datetime.datetime.strptime(request.POST.get('fecha_fin'), '%Y-%m-%d')
            
            Restriccion.objects.all().delete()  # eliminar las restricciones anteriores
            restriccion = Restriccion(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
            restriccion.save()  # guardar la nueva restricción
            
            messages.success(request, f'Se ha restringido el ingreso de disponibilidad para el rango de fechas seleccionado.')
            return redirect('/inicio/seleccionar-fechas')
    
    return render(request, 'restringirprueba.html', {'fecha_actual': fecha_actual})


import datetime

def vista_para_profesor(request):
    fecha_actual = datetime.datetime.now().strftime('%Y-%m-%d')
    form = DisponibilidadForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        # obtener los datos del formulario y el usuario actual
        fecha = form.cleaned_data['fecha']
        hora_inicio = form.cleaned_data['hora_inicio']
        hora_fin = form.cleaned_data['hora_fin']
        profesor = request.user

        # verificar si el día está restringido
        restricciones = Restriccion.objects.all()
        for restriccion in restricciones:
            if restriccion.fecha_inicio <= datetime.datetime.strptime(fecha_actual, '%Y-%m-%d').date() <= restriccion.fecha_fin:
                messages.error(request, f"No puedes hacer disponibilidad en este día porque ha sido restringido entre las {restriccion.fecha_inicio} y las {restriccion.fecha_fin}.")
                return redirect('/inicio/disponibilidad')
        
        # verificar si ya hay disponibilidad en el mismo horario
        disponibilidades = Disponibilidad.objects.filter(
            profesor=profesor,
            fecha=fecha,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin
        )
        if disponibilidades.exists():
            messages.error(request, 'Ya tienes una disponibilidad registrada para este horario.')
            return redirect('/inicio/disponibilidad')

        # guardar la disponibilidad en la base de datos
        disponibilidad = Disponibilidad(
            fecha=fecha,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
            comentarios=form.cleaned_data['comentarios'],
            profesor=profesor,
        )
        disponibilidad.save()
        messages.success(request, 'Se ha guardado la disponibilidad correctamente.')
        return redirect('/inicio/disponibilidad')

    return render(request, 'disponibilidad.html', {'fecha_actual': fecha_actual, 'form': form})



def vereditdisponibilidad(request):
    ordenar_por = request.GET.get('ordenar_por', 'fecha') # por defecto, ordenar por fecha
    disponibilidad = Disponibilidad.objects.order_by(ordenar_por)
    context = {
        'disponibilidad': disponibilidad,
        'ordenar_por': ordenar_por,
    }
    return render(request, 'editardisponibilidad.html', context)

def activar_disponibilidad(request, disponibilidad_id):
    disponibilidad = get_object_or_404(Disponibilidad, id=disponibilidad_id)
    disponibilidad.mostrar_en_tabla = True 
    disponibilidad.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))

def desactivar_disponibilidad(request, disponibilidad_id):
    disponibilidad = get_object_or_404(Disponibilidad, id=disponibilidad_id)
    disponibilidad.mostrar_en_tabla = False
    disponibilidad.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))


def editar_disponibilidad(request, disponibilidad_id):
    disponibilidad = get_object_or_404(Disponibilidad, id=disponibilidad_id)

    if request.method == 'POST':
        form = EditDisponibilidadForm(request.POST, instance=disponibilidad)
        if form.is_valid():
            form.save()
            return redirect('/inicio/editarDisponibilidad')
    else:
        form = EditDisponibilidadForm(instance=disponibilidad)

    return render(request, 'editarDisponibilidad.html', {'form': form})
 

def ver_restricciones(request):
    restricciones = Restriccion.objects.all()
    return render(request, 'verRestricciones.html', {'restricciones': restricciones})


def eliminar_restriccion(request, restriccion_id):
    # Obtener la restricción a eliminar
    restriccion = Restriccion.objects.get(pk=restriccion_id)

    if request.method == 'POST':
        # Confirmar que se quiere eliminar la restricción
        restriccion.delete()
        messages.success(request, 'La restricción ha sido eliminada correctamente.')
        return redirect('inicio/verRestricciones')

    return render(request, 'verRestricciones.html', {'restriccion': restriccion})


from django.contrib.auth.models import User
from django.db.models import Q
def cargar_tabla(request):
    if request.method == 'POST':
        archivo = request.FILES['archivo_excel']
        workbook = openpyxl.load_workbook(archivo)  # Cargar el archivo Excel con openpyxl
        
        # Obtener los datos de las columnas y guardarlos en las tablas correspondientes
        datos_tabla1 = []
        
        sheet = workbook['programación']  # Reemplaza 'NombreHoja' con el nombre de la hoja en tu archivo Excel
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            programa = row[0]
            prog_jornada = row[1]
            tipo_asignatura = row[2]
            semestre = row[3]
            codigo_asig = row[4]
            nombre_asig = row[5]
            creditos = row[6]
            grupo = row[7]
            codigo_grupo = row[8]
            cupo = row[9]
            cupo_generico = row[10]
            ajuste_cupos = row[11]
            cc = row[12]
            nombre_docente = row[13]
            correo_electronico = row[14]
            dia = row[15]
            horario = row[16]
            intensidad_semanal = row[17]
            clase_requiere_sala = row[18]
            salon = row[19]
            observaciones = row[20]
            banco_datos_docentes = row[21]
            salones = row[22]
            capacidad_espacios = row[23]
            print("Datos de la fila:")
            print("Programa:", programa)
            print("Programa jornada:", prog_jornada)
            print("Tipo de asignatura:", tipo_asignatura)
            print("Semestre:", semestre)
            print("Código de asignatura:", codigo_asig)
            print("Nombre de asignatura:", nombre_asig)
            print("Créditos:", creditos)
            print("Grupo:", grupo)
            print("Código de grupo:", codigo_grupo)
            print("Cupo:", cupo)
            print("Cupo genérico:", cupo_generico)
            print("Ajuste de cupos:", ajuste_cupos)
            print("CC:", cc)
            print("Nombre de docente:", nombre_docente)
            print("Correo electrónico:", correo_electronico)
            print("Día:", dia)
            print("Horario:", horario)
            print("Intensidad semanal:", intensidad_semanal)
            print("Clase requiere sala:", clase_requiere_sala)
            print("Salón:", salon)
            print("Observaciones:", observaciones)
            print("Banco de datos docentes:", banco_datos_docentes)
            print("Salones:", salones)
            print("Capacidad de espacios:", capacidad_espacios)
            

            cc = str(cc)
            nombres = nombre_docente.split()  # Separar los nombres por espacios
            primer_apellidos = nombres[0]  # Obtener el primer nombre
            segundo_apellidos = nombres[1]  # Obtener el segundo nombre
            nombres = nombres[2:]  # Obtener los apellidos como una lista

            # Crear usuarios y guardarlos en la tabla User
            user, created = User.objects.get_or_create(username=cc, email=correo_electronico)
            user.first_name = ' '.join(nombres)
            user.last_name = f"{ primer_apellidos} {segundo_apellidos}"  # Convertir la lista de apellidos en una cadena separada por espacios
            user.set_password(cc)
            user.save()
            grupo_profesor, created = Group.objects.get_or_create(name='Profesores')  
            user.groups.add(grupo_profesor)
           
            try:
                # Verificar si la asignatura ya existe
                asignatura = Asignatura.objects.get(codigo=codigo_asig)
            except Asignatura.DoesNotExist:
                    # La asignatura no existe, crear una nueva
                    asignatura = Asignatura(
                        codigo=codigo_asig,
                        nombre=nombre_asig,
                        creditos=creditos,
                        intensidad=creditos * 2
                    )
            asignatura.save()
            try:
                # Verificar si el salon ya existe
                salon = Salones.objects.get(nombre=salon)
            except Salones.DoesNotExist:
                    # el salon no existe, crear uno nuevo
                    salon = Salones(
                        nombre=salon,
                        tipo=salones,
                        capacidad=capacidad_espacios,
                        
                    )
            salon.save()
            
            programacion = Programacion.objects.filter(
                Q(programa_jornada=prog_jornada) &
                Q(codigo_asignatura=codigo_asig) &
                Q(grupo=grupo) &
                Q(codigo_grupo=codigo_grupo) &
                Q(cupo=cupo) &
                Q(cupo_generico=cupo_generico) &
                Q(salon=salon) &
                Q(id_usuarios=user)
            ).exists()
            
            
            if programacion:
        # La programación ya existe, realiza la acción necesaria
        
                print("Error: La programación ya existe.")
            else:
             programacion = Programacion(
                    programa_jornada=prog_jornada,
                    codigo_asignatura=codigo_asig,
                    grupo=grupo,
                    codigo_grupo=codigo_grupo,
                    cupo=cupo,
                    cupo_generico=cupo_generico,
                    salon=salon,
                    id_usuarios=user
                    
                )
             programacion.save()
            
            
        
        return render(request, 'cargarTabla.html')
    
    return render(request, 'cargarTabla.html')



def mostrar_programacion(request):
    programacion = Programacion.objects.all()
    return render(request, 'programacion.html', {'programacion': programacion})





def mostrar_cronograma(request):
    if request.method == 'POST':
        profesor_id = request.POST.get('profesor_id')
        profesor = get_object_or_404(User.objects.filter(groups__name='Profesores'), id=profesor_id)
        # Obtener todas las programaciones relacionadas con el profesor seleccionado
        programaciones = Programacion.objects.filter(id_usuarios=profesor)
        # Obtener los códigos de asignatura de las programaciones
        codigos_asignatura = programaciones.values_list('codigo_asignatura', flat=True)
        # Obtener los códigos del programa de las programaciones
        codigos_programas = programaciones.values_list('programa_jornada', flat=True)
        # Obtener todos los profesores disponibles para mostrar en el formulario
        profesores = User.objects.filter(groups__name='Profesores')
        asignaturas = Asignatura.objects.filter(codigo__in=codigos_asignatura)
        programas = Programas.objects.filter(cod__in=codigos_programas)
        # Obtener los datos del cronograma relacionados con las programaciones
        cronogramas = Cronograma.objects.filter(id_usuarios=profesor)
        # Pasar los datos a la plantilla
        context = {
            'profesor_seleccionado': profesor,
            'codigos_asignatura': codigos_asignatura,
            'programaciones': programaciones,
            'profesores': profesores,
            'asignaturas': asignaturas,
            'codigos_programas': codigos_programas,
            'programas': programas,
            'cronogramas': cronogramas,
        }
        return render(request, 'mostrarCronograma.html', context)

    profesores = User.objects.filter(groups__name='Profesores')
    context = {
        'profesores': profesores,
    }
    return render(request, 'mostrarCronograma.html', context)






def llenar_cronograma(request):
    if request.method == 'POST':
        form = CronogramaForm(request.POST, user=request.user)  # Pasar el usuario actual al formulario
        if form.is_valid():
            # Procesar los datos del formulario
            semana = form.cleaned_data['semana']
            fecha = form.cleaned_data['fecha']
            contenido_tematico = form.cleaned_data['contenido_tematico']
            material_apoyo = form.cleaned_data['material_apoyo']
            observaciones = form.cleaned_data['observaciones']
            chequeo = form.cleaned_data['chequeo']

            cronograma = Cronograma(
                id_usuarios=request.user,
                semana=semana,
                fecha=fecha,
                contenido_tematico=contenido_tematico,
                material_apoyo=material_apoyo,
                observaciones=observaciones,
                chequeo=chequeo
            )
            cronograma.save()
            messages.success(request, '¡Los datos se han guardado exitosamente!')

            # Redireccionar o mostrar un mensaje de éxito
         
            return redirect('/inicio/cronograma')

    else:
        initial_data = {'profesor_id': request.user}  # Establecer el valor inicial del campo profesor_id
        form = CronogramaForm(user=request.user, initial=initial_data)  # Pasar el usuario actual y los datos iniciales al formulario
       

    return render(request, 'llenarCronograma.html', {'form': form})


def mostrar_cronograma_a_profesor(request):
    # Recuperar todos los objetos Cronograma de la base de datos
    cronogramas = Cronograma.objects.all()

    return render(request, 'mostrarCronogramaAprofesor.html', {'cronogramas': cronogramas})



def editar_cronograma(request, cronograma_id):
    cronograma = get_object_or_404(Cronograma, pk=cronograma_id)
    
    if request.method == 'POST':
        form = EditCronogramaForm(request.POST)
        if form.is_valid():
            cronograma.semana = form.cleaned_data['semana']
            cronograma.fecha = form.cleaned_data['fecha']
            cronograma.contenido_tematico = form.cleaned_data['contenido_tematico']
            cronograma.material_apoyo = form.cleaned_data['material_apoyo']
            cronograma.observaciones = form.cleaned_data['observaciones']
            cronograma.chequeo = form.cleaned_data['chequeo']
            cronograma.save()
            # Redirigir a una página de éxito o realizar otras acciones
            return redirect('/inicio/cronograma')
            
    else:
        
        form = EditCronogramaForm(initial={
            'semana': cronograma.semana,
            'fecha': cronograma.fecha,
            'contenido_tematico': cronograma.contenido_tematico,
            'material_apoyo': cronograma.material_apoyo,
            'observaciones': cronograma.observaciones,
            'chequeo': cronograma.chequeo
        })
    
    context = {
        'form': form
    }
    
    return render(request, 'editarCronograma.html', context)

def editar_cronograma1(request, cronograma_id):
    cronograma = get_object_or_404(Cronograma, pk=cronograma_id)
    
    if request.method == 'POST':
        form = EditCronogramaForm1(request.POST)
        if form.is_valid():
            cronograma.semana = form.cleaned_data['semana']
            cronograma.fecha = form.cleaned_data['fecha']
            cronograma.contenido_tematico = form.cleaned_data['contenido_tematico']
            cronograma.material_apoyo = form.cleaned_data['material_apoyo']
            cronograma.observaciones = form.cleaned_data['observaciones']
            cronograma.chequeo = form.cleaned_data['chequeo']
            cronograma.save()
            # Redirigir a una página de éxito o realizar otras acciones
            return redirect('/inicio/mostrar_cronograma')
            
    else:
        form = EditCronogramaForm1(initial={
            'semana': cronograma.semana,
            'fecha': cronograma.fecha,
            'contenido_tematico': cronograma.contenido_tematico,
            'material_apoyo': cronograma.material_apoyo,
            'observaciones': cronograma.observaciones,
            'chequeo': cronograma.chequeo
        })
    
    context = {
        'form': form
    }
    
    return render(request, 'editarCronograma.html', context)


def activar_cronograma(request, cronograma_id):
    cronograma = get_object_or_404(Cronograma, id=cronograma_id)
    cronograma.mostrar_en_tabla = True 
    cronograma.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))

def desactivar_cronograma(request, cronograma_id):
    cronograma = get_object_or_404(Cronograma, id=cronograma_id)
    cronograma.mostrar_en_tabla = False
    cronograma.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))


from django.shortcuts import render
from django.db.models import Count
from .models import Cronograma
from django.contrib.auth.models import User, Group

def contador_registros(request):
    # Obtén el grupo "profesor"
    grupo_profesor = Group.objects.get(name='Profesores')

    # Obtén todos los usuarios del grupo "profesor"
    usuarios_profesor = User.objects.filter(groups=grupo_profesor)

    # Realiza la agregación para contar el número de registros por usuario
    contador_registros = Cronograma.objects.filter(id_usuarios__in=usuarios_profesor).values('id_usuarios').annotate(total_registros=Count('id'))

    # Crea un diccionario para almacenar el número de registros y el nombre de usuario por usuario
    registros_por_usuario = {usuario.id: {'username': usuario.username, 'total_registros': 0} for usuario in usuarios_profesor}

    # Actualiza el diccionario con el número de registros contados
    for contador in contador_registros:
        usuario_id = contador['id_usuarios']
        total_registros = contador['total_registros']
        registros_por_usuario[usuario_id]['total_registros'] = total_registros

    # Pasa los datos a la plantilla para su visualización
    context = {
        'registros_por_usuario': registros_por_usuario,
    }

    return render(request, 'contador_registros.html', context)


def verCronograma(request, profesor_id):
    try:
        profesor = User.objects.get(id=profesor_id)
        cronogramas = Cronograma.objects.filter(id_usuarios=profesor)
    except User.DoesNotExist:
        cronogramas = None

    context = {
        'profesor': profesor,
        'cronogramas': cronogramas,
    }

    return render(request, 'verCronograma.html', context)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.http import HttpResponse
from .models import Cronograma  # Importa el modelo de los cronogramas
from django.utils import timezone

def descargar_reporte_excel(request, id_usuario):
    # Obtén los cronogramas del usuario especificado por ID
    cronogramas = Cronograma.objects.filter(id_usuarios =id_usuario)

    # Crea un nuevo libro de trabajo de Excel
    wb = Workbook()
    sheet = wb.active

    # Agrega los encabezados de columna
    encabezados = ['Username','Semana', 'Fecha', 'Contenido Temático', 'Material de Apoyo', 'Observaciones']
    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = get_column_letter(col_num)
        sheet[f'{col_letra}1'] = encabezado
        sheet[f'{col_letra}1'].alignment = Alignment(horizontal='center')

    # Agrega los datos de los cronogramas al libro de trabajo
    row_index = 2
    for cronograma in cronogramas:
        sheet[f'A{row_index}'] = cronograma.id_usuarios.username
        sheet[f'B{row_index}'] = cronograma.semana
        sheet[f'C{row_index}'] = cronograma.fecha.replace(tzinfo=None) if cronograma.fecha else None
        sheet[f'D{row_index}'] = cronograma.contenido_tematico
        sheet[f'E{row_index}'] = cronograma.material_apoyo
        sheet[f'F{row_index}'] = cronograma.observaciones
        row_index += 1

    # Ajusta el ancho de las columnas automáticamente
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length

    # Guarda el libro de trabajo en un objeto de tipo HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte.xlsx"'
    wb.save(response)

    return response

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

def descargar_reporte_pdf(request, id_usuario):
    # Obtén los cronogramas del usuario especificado por ID
    cronogramas = Cronograma.objects.filter(id_usuarios=id_usuario)

    # Crea un nuevo objeto de lienzo de PDF
    response = HttpResponse(content_type='application/pdf')
    
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'

    # Crea el documento PDF con el tamaño de página letter
    doc = SimpleDocTemplate(response, pagesize=letter)

    # Estilos de celda para el encabezado y los datos
    estilo_celda_encabezado = [('BACKGROUND', (0, 0), (-1, 0), colors.red)]
    estilo_celda_datos = [('GRID', (0, 0), (-1, -1), 1, colors.black)]

    # Crea la tabla con los datos de los cronogramas
    data = [['Profesor', 'Semana', 'Fecha', 'Contenido Temático', 'Material de Apoyo', 'Observaciones']]
    for cronograma in cronogramas:
        data.append([
            cronograma.id_usuarios.username,
            cronograma.semana,
            cronograma.fecha.strftime('%Y-%m-%d') if cronograma.fecha else None,
            cronograma.contenido_tematico,
            cronograma.material_apoyo,
            cronograma.observaciones
        ])

    tabla = Table(data)
    tabla.setStyle(TableStyle(estilo_celda_encabezado + estilo_celda_datos))

    # Construye el documento PDF
    elementos = [tabla]
    doc.build(elementos)

    return response

from django.shortcuts import render
from django.contrib.auth.models import User
from .models import Programas, Programacion, Cronograma

def programaciones_asignadas_view(request):
    programa_tecnico = None
    programaciones_asignadas = None
    cronogramas_asignados = None

    if request.user.groups.filter(name='Tecnicos de apoyo').exists():
        tecnico_apoyo = request.user.programas.first()
        if tecnico_apoyo:
            programa_tecnico = tecnico_apoyo.cod
            programaciones_asignadas = Programacion.objects.filter(programa_jornada=programa_tecnico)
            
            if request.method == 'POST':
                usuario_id = request.POST.get('usuario_id')
                id_usuarios = programaciones_asignadas.values_list('id_usuarios', flat=True)
                if usuario_id == 'todos':
                    cronogramas_asignados = Cronograma.objects.filter(id_usuarios__in=id_usuarios)
                else:
                    cronogramas_asignados = Cronograma.objects.filter(id_usuarios__username=usuario_id)
    programas_asignados = Programas.objects.filter(tecnico_apoyo=request.user)

    context = {
        'programas_asignados': programas_asignados,
        'programa_tecnico': programa_tecnico,
        'programaciones_asignadas': programaciones_asignadas,
        'cronogramas_asignados': cronogramas_asignados,
    }
    return render(request, 'programaciones_asignadas.html', context)



def descargar_reporte_view(request, id_usuario):
    if id_usuario == 'todos':
        # Obtén todos los cronogramas
        cronogramas = Cronograma.objects.all()
    else:
        # Verifica si se seleccionó un usuario específico o el valor es inválido
        try:
            usuario = User.objects.get(id=id_usuario)
            cronogramas = Cronograma.objects.filter(id_usuarios=usuario)
        except User.DoesNotExist:
            cronogramas = None

    if cronogramas is None:
        # Manejo de error: usuario inválido
        return HttpResponse('Usuario inválido.')

    # Crea un nuevo libro de trabajo de Excel
    wb = Workbook()
    sheet = wb.active

    # Agrega los encabezados de columna
    encabezados = ['Profesor', 'Semana', 'Fecha', 'Contenido Temático', 'Material de Apoyo', 'Observaciones']
    for col_num, encabezado in enumerate(encabezados, 1):
        col_letra = get_column_letter(col_num)
        sheet[f'{col_letra}1'] = encabezado
        sheet[f'{col_letra}1'].alignment = Alignment(horizontal='center')

    # Agrega los datos de los cronogramas al libro de trabajo
    row_index = 2
    for cronograma in cronogramas:
        sheet[f'A{row_index}'] = cronograma.id_usuarios.username
        sheet[f'B{row_index}'] = cronograma.semana
        sheet[f'C{row_index}'] = cronograma.fecha.strftime('%Y-%m-%d') if cronograma.fecha else None
        sheet[f'D{row_index}'] = cronograma.contenido_tematico
        sheet[f'E{row_index}'] = cronograma.material_apoyo
        sheet[f'F{row_index}'] = cronograma.observaciones
        row_index += 1

    # Ajusta el ancho de las columnas automáticamente
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length

    # Guarda el libro de trabajo en un objeto de tipo HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte.xlsx"'
    wb.save(response)

    return response
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

def descargar_reporte_viewpdf(request, id_usuario):
    if id_usuario == 'todos':
        # Obtén todos los cronogramas
        cronogramas = Cronograma.objects.all()
    else:
        # Obtén los cronogramas del usuario especificado por ID
        cronogramas = Cronograma.objects.filter(id_usuarios=id_usuario)

    # Crea un documento PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'

    # Configura el documento PDF
    doc = SimpleDocTemplate(response, pagesize=letter)

    # Define los estilos de la tabla
    estilo_celda_encabezado = [('BACKGROUND', (0, 0), (-1, 0), colors.red)]
    estilo_celda_datos = [('GRID', (0, 0), (-1, -1), 1, colors.black)]

    # Crea la tabla con los datos de los cronogramas
    data = [['Profesor', 'Semana', 'Fecha', 'Contenido Temático', 'Material de Apoyo', 'Observaciones']]
    for cronograma in cronogramas:
        data.append([
            cronograma.id_usuarios.username,
            cronograma.semana,
            cronograma.fecha.strftime('%Y-%m-%d') if cronograma.fecha else None,
            cronograma.contenido_tematico,
            cronograma.material_apoyo,
            cronograma.observaciones
        ])

    tabla = Table(data)
    tabla.setStyle(TableStyle(estilo_celda_encabezado + estilo_celda_datos))

    # Construye el documento PDF
    elementos = [tabla]
    doc.build(elementos)

    return response





from .models import Asistencia
from django.shortcuts import render, redirect

def asistencia(request):
    usuarios = User.objects.filter(groups__name='Profesores')
    semanas = range(1, 17)
    usuario_seleccionado = request.GET.get('usuario')
    semana_seleccionada = request.GET.get('semana')

    cronogramas_asignados = Cronograma.objects.filter(id_usuarios__username=usuario_seleccionado, semana=semana_seleccionada)

    if request.method == 'POST':
        asistencias = request.POST.getlist('asistencia')
        fecha_recuperacion = request.POST.get('fecha_recuperacion')
        tema_clase = request.POST.get('tema_clase')
        
        # Obtener el objeto User correspondiente al usuario seleccionado
        usuario = User.objects.get(username=usuario_seleccionado)

        for asistencia_value in asistencias:
            cronograma_id, asistencia_type = asistencia_value.split('-')

            cronograma = Cronograma.objects.get(id=cronograma_id)

            # Obtener el objeto Programacion correspondiente al usuario y asignar el salón
            programacion = Programacion.objects.get(id_usuarios=usuario)
            salon_asignado = programacion.salon

            if asistencia_type == 'noasistio' and fecha_recuperacion:
                # Crear el objeto Asistencia con fecha de recuperación y asignar el salón
                asistencia = Asistencia(
                    cronograma=cronograma,
                    usuario=usuario,
                    fecha=cronograma.fecha,
                    asistio=False,
                    noAsistio=True,
                    fecha_recuperacion=fecha_recuperacion,
                    
                )
            else:
                # Crear el objeto Asistencia sin fecha de recuperación y asignar el salón
                asistencia = Asistencia(
                    cronograma=cronograma,
                    usuario=usuario,
                    fecha=cronograma.fecha,
                    asistio=(asistencia_type == 'asistio'),
                    noAsistio=(asistencia_type == 'noasistio'),
                    tema_clase=tema_clase,
                    salon=salon_asignado
                )

            # Guardar la asistencia
            asistencia.save()

        return redirect('asistencia')  # Redirigir a la página de asistencia después de guardar las asistencias

    return render(request, 'asistencia.html', {
        'usuarios': usuarios,
        'semanas': semanas,
        'usuario_seleccionado': usuario_seleccionado,
        'semana_seleccionada': semana_seleccionada,
        'cronogramas_asignados': cronogramas_asignados
    })
    
    

def tabla_asistencia(request):
    asistencias = Asistencia.objects.filter(salon__isnull=True).exclude(fecha_recuperacion__isnull=True)
    todos_los_salones = Salones.objects.all()

    if request.method == 'POST':
        usuario = request.POST.get('usuario')
        if usuario:
            asistencias = asistencias.filter(usuario=usuario)

    return render(request, 'tabla_asistencia.html', {'asistencias': asistencias, 'salones': todos_los_salones, 'usuarios': asistencias.values_list('usuario', flat=True).distinct()})

def guardar_salon(request):
    if request.method == 'POST':
        asistencia_id = request.POST.get('asistencia_id')
        salon_id = request.POST.get('salon')

        asistencia = Asistencia.objects.get(id=asistencia_id)
        salon = Salones.objects.get(id=salon_id)
        asistencia.salon = salon
        asistencia.save()

        asistencias = Asistencia.objects.filter(salon__isnull=True).exclude(fecha_recuperacion__isnull=True).exclude(id=asistencia_id)
        return render(request, 'tabla_asistencia.html', {'asistencias': asistencias, 'salones': Salones.objects.all(), 'usuarios': asistencias.values_list('usuario', flat=True).distinct()})

    return redirect('tabla_asistencia')


def mostrar_tabla_asistencias(request):
    asistencias = Asistencia.objects.all()
    usuarios_asistencias = asistencias.values_list('usuario__username', flat=True).distinct()
    usuarios = User.objects.filter(username__in=usuarios_asistencias)

    usuario_seleccionado = request.POST.get('usuario')

    if usuario_seleccionado:
        asistencias = asistencias.filter(usuario__username=usuario_seleccionado)

    context = {'asistencias': asistencias, 'usuarios': usuarios, 'usuario_seleccionado': usuario_seleccionado}
    return render(request, 'mostrarAsistencia.html', context)

def mostrar_tabla_asistencias_profesor(request):
    usuario = request.user  # Obtener el usuario que inició sesión
    asistencias = Asistencia.objects.filter(usuario=usuario)
  

    usuario_seleccionado = request.POST.get('usuario')

    if usuario_seleccionado:
        asistencias = asistencias.filter(usuario__username=usuario_seleccionado)

    context = {'asistencias': asistencias}
    return render(request, 'mostrarAsistenciaProfesor.html', context)






def asistenciaProfesor(request):
    # Obtener el usuario que inició sesión
    usuario_actual = request.user

    semanas = range(1, 17)
    semana_seleccionada = request.GET.get('semana')

    cronogramas_asignados = Cronograma.objects.filter(id_usuarios__username=usuario_actual.username, semana=semana_seleccionada)

    if request.method == 'POST':
            asistencias = request.POST.getlist('asistencia')
            fecha_recuperacion = request.POST.get('fecha_recuperacion')
            tema_clase = request.POST.get('tema_clase')

            for asistencia_value in asistencias:
                cronograma_id, asistencia_type = asistencia_value.split('-')

                cronograma = Cronograma.objects.get(id=cronograma_id)

                # Obtener el objeto Programacion correspondiente al usuario y asignar el salón
                programacion = Programacion.objects.get(id_usuarios=usuario_actual)
                salon_asignado = programacion.salon

                if asistencia_type == 'noasistio' and fecha_recuperacion:
                    # Crear el objeto Asistencia con fecha de recuperación y asignar el salón
                    asistencia = Asistencia(
                        cronograma=cronograma,
                        usuario=usuario_actual,
                        fecha=cronograma.fecha,
                        asistio=False,
                        noAsistio=True,
                        fecha_recuperacion=fecha_recuperacion,
                    )
                else:
                    # Crear el objeto Asistencia sin fecha de recuperación y asignar el salón
                    asistencia = Asistencia(
                        cronograma=cronograma,
                        usuario=usuario_actual,
                        fecha=cronograma.fecha,
                        asistio=(asistencia_type == 'asistio'),
                        noAsistio=(asistencia_type == 'noasistio'),
                        tema_clase=tema_clase,
                        salon=salon_asignado
                    )

                # Guardar la asistencia
                asistencia.save()

            return redirect('asistenciaProfesor')  # Redirigir a la página de asistencia después de guardar las asistencias

    return render(request, 'asistenciaProfesor.html', {
            'semanas': semanas,
            'semana_seleccionada': semana_seleccionada,
            'cronogramas_asignados': cronogramas_asignados
        })
    
